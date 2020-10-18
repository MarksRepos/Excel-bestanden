import os
from openpyxl import load_workbook, Workbook

print(os.getcwd())

wb = load_workbook(filename='voorbeeld.xlsx')

ws = wb['Personen']

# test
# for cell in ws['B']:
#     print(cell.value)

# for rij in ws['A1:D11']:
#     for cell in rij:
#         print(cell.value)

# Begin bij row 2 (zonder header)
# for rij in ws.iter_rows(min_row=2):
#     print(rij)

class Persoon():

    def __init__(self, naam, gewicht, leeftijd, woonplaats):
        self.naam = naam
        self.gewicht = gewicht
        self.leeftijd = leeftijd
        self.woonplaats = woonplaats

    def is_jarig(self):
        self.leeftijd = self.leeftijd + 1
        print(f'{self.naam} is nu {self.leeftijd} jaren oud. Hoera')


personen = []

for rij in ws.iter_rows(min_row=2):

    persoon = Persoon(rij[0].value, rij[1].value, rij[2].value, rij[3].value)
    persoon.is_jarig()
    personen.append(persoon)

wb = Workbook()
ws = wb.active

for persoon in personen:
    ws.append([persoon.naam, persoon.gewicht, persoon.leeftijd, persoon.woonplaats])

wb.save('een jaartje ouder.xlsx')

print("\nKlaar!")



